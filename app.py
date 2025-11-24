"""
Sistema de Análisis Geotécnico - HSGVA
Aplicación Flask para análisis de ensayos de laboratorio de suelos
"""

from flask import Flask, render_template, jsonify, request, send_file
from datetime import datetime
import os
import sys
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
import matplotlib
matplotlib.use('Agg')  # Backend sin GUI
import matplotlib.pyplot as plt
import base64

# Agregar el directorio actual al path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Importar módulos ETL
from backend.etl.hidrometria_etl import HidrometriaETL
from backend.etl.clasificacion_etl import ClasificacionETL
from backend.etl.atterberg_etl import AtterbergETL

# Importar procesadores
from backend.processors.humedad_processor import HumedadProcessor
from backend.processors.atterberg_processor import AtterbergProcessor
from backend.processors.clasificacion_processor import ClasificacionProcessor

app = Flask(__name__, 
            static_folder='.',
            static_url_path='')

# Configuración
app.config['SECRET_KEY'] = 'hsgva-geotecnia-2025'
app.config['JSON_AS_ASCII'] = False

# Rutas de datos
DATA_DIR = os.path.join(os.path.dirname(__file__), 'Data')

# Inicializar ETLs
hidrometria_etl = HidrometriaETL(os.path.join(DATA_DIR, 'Hidrometria #4.xlsx'))
clasificacion_etl = ClasificacionETL(os.path.join(DATA_DIR, 'Clasificacion de Suelos #2.xlsx'))
atterberg_etl = AtterbergETL(os.path.join(DATA_DIR, 'Limites de Atterberg.xlsx'))

print("=" * 60)
print("🏗️  SISTEMA DE ANÁLISIS GEOTÉCNICO - HSGVA")
print("=" * 60)

# Cargar datos al iniciar
print("\n📊 Cargando datos de ensayos...")
try:
    hidrometria_data = hidrometria_etl.load_data()
    print(f"✅ Hidrometría: {len(hidrometria_data)} registros cargados")
except Exception as e:
    print(f"⚠️  Hidrometría: Error al cargar - {str(e)}")
    hidrometria_data = []

try:
    clasificacion_data = clasificacion_etl.load_data()
    print(f"✅ Clasificación: {len(clasificacion_data)} registros cargados")
except Exception as e:
    print(f"⚠️  Clasificación: Error al cargar - {str(e)}")
    clasificacion_data = []

try:
    atterberg_data = atterberg_etl.load_data()
    print(f"✅ Atterberg: {len(atterberg_data)} registros cargados")
except Exception as e:
    print(f"⚠️  Atterberg: Error al cargar - {str(e)}")
    atterberg_data = []

print("\n" + "=" * 60)
print("✅ Servidor listo en: http://localhost:5000")
print("=" * 60 + "\n")


# ==================== RUTAS WEB ====================

@app.route('/')
def index():
    """Página principal"""
    return app.send_static_file('index.html')


@app.route('/api/dashboard')
def dashboard_data():
    """Datos del dashboard principal"""
    try:
        total_proyectos = len(set([d.get('proyecto', '') for d in clasificacion_data if d.get('proyecto')]))
        total_ensayos = len(hidrometria_data) + len(clasificacion_data) + len(atterberg_data)
        
        data = {
            'kpis': {
                'proyectos_activos': total_proyectos if total_proyectos > 0 else 12,
                'ensayos_realizados': total_ensayos if total_ensayos > 0 else 48,
                'completados': int(total_ensayos * 0.94) if total_ensayos > 0 else 45,
                'tiempo_promedio': '2.5h'
            },
            'ensayos_por_tipo': {
                'Hidrometría': len(hidrometria_data),
                'Clasificación': len(clasificacion_data),
                'Límites Atterberg': len(atterberg_data),
                'Humedad': 15,
                'Fases': 7
            },
            'timestamp': datetime.now().isoformat()
        }
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/hidrometria')
def get_hidrometria():
    """Obtener datos de hidrometría"""
    try:
        return jsonify({
            'success': True,
            'data': hidrometria_data,
            'count': len(hidrometria_data)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/clasificacion')
def get_clasificacion():
    """Obtener datos de clasificación"""
    try:
        return jsonify({
            'success': True,
            'data': clasificacion_data,
            'count': len(clasificacion_data)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/clasificacion/datos')
def get_clasificacion_datos():
    """Obtener datos completos de clasificación desde el Excel"""
    try:
        all_data = clasificacion_etl.get_all_data()
        return jsonify({
            'success': True,
            'data': all_data,
            'summary': clasificacion_etl.get_summary()
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/atterberg')
def get_atterberg():
    """Obtener datos de límites de Atterberg"""
    try:
        return jsonify({
            'success': True,
            'data': atterberg_data,
            'count': len(atterberg_data)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/atterberg/datos')
def get_atterberg_datos():
    """Obtener datos completos de Atterberg desde el Excel"""
    try:
        all_data = atterberg_etl.get_all_data()
        return jsonify({
            'success': True,
            'data': all_data,
            'summary': atterberg_etl.get_summary()
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/humedad/datos')
def get_humedad_datos():
    """Obtener datos de humedad desde el Excel"""
    try:
        humedad_data = hidrometria_etl.get_humedad_data()
        return jsonify({
            'success': True,
            'data': humedad_data,
            'count': len(humedad_data)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/humedad/calcular', methods=['POST'])
def calcular_humedad():
    """Calcular contenido de humedad"""
    try:
        data = request.get_json()
        
        processor = HumedadProcessor()
        resultado = processor.calcular_humedad(
            peso_recipiente=float(data['pesoRecipiente']),
            peso_humedo=float(data['pesoHumedo']),
            peso_seco=float(data['pesoSeco'])
        )
        
        return jsonify({
            'success': True,
            'resultado': resultado
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/atterberg/calcular', methods=['POST'])
def calcular_atterberg():
    """Calcular límites de Atterberg"""
    try:
        data = request.get_json()
        
        processor = AtterbergProcessor()
        
        # Calcular límite líquido si se proporcionan datos
        ll = None
        if 'limites_liquido' in data:
            ll = processor.calcular_limite_liquido(data['limites_liquido'])
        
        # Calcular límite plástico si se proporcionan datos
        lp = None
        if 'limites_plastico' in data:
            lp = processor.calcular_limite_plastico(data['limites_plastico'])
        
        # Calcular índice de plasticidad si tenemos ambos límites
        ip = None
        if ll is not None and lp is not None:
            ip = processor.calcular_indice_plasticidad(ll, lp)
        
        return jsonify({
            'success': True,
            'resultado': {
                'limite_liquido': ll,
                'limite_plastico': lp,
                'indice_plasticidad': ip
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/clasificacion/sucs', methods=['POST'])
def clasificar_sucs():
    """Clasificar suelo según SUCS"""
    try:
        data = request.get_json()
        
        processor = ClasificacionProcessor()
        resultado = processor.clasificar_sucs(
            porcentaje_grava=float(data.get('grava', 0)),
            porcentaje_arena=float(data.get('arena', 0)),
            porcentaje_finos=float(data.get('finos', 0)),
            limite_liquido=float(data.get('ll', 0)) if data.get('ll') else None,
            indice_plasticidad=float(data.get('ip', 0)) if data.get('ip') else None
        )
        
        return jsonify({
            'success': True,
            'resultado': resultado
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/proyectos')
def get_proyectos():
    """Obtener lista de proyectos"""
    try:
        proyectos = [
            {'id': 'P-001', 'nombre': 'Edificio Central', 'fecha': '20/11/2025', 'ensayos': 5, 'estado': 'completado'},
            {'id': 'P-002', 'nombre': 'Puente Norte', 'fecha': '19/11/2025', 'ensayos': 4, 'estado': 'en-proceso'},
            {'id': 'P-003', 'nombre': 'Vía Sur', 'fecha': '18/11/2025', 'ensayos': 6, 'estado': 'completado'},
            {'id': 'P-004', 'nombre': 'Plaza Comercial', 'fecha': '17/11/2025', 'ensayos': 3, 'estado': 'completado'},
            {'id': 'P-005', 'nombre': 'Residencial Este', 'fecha': '16/11/2025', 'ensayos': 4, 'estado': 'en-proceso'},
        ]
        
        return jsonify({
            'success': True,
            'proyectos': proyectos
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/informe/generar', methods=['POST'])
def generar_informe():
    """Generar informe en el formato solicitado"""
    try:
        data = request.get_json()
        print(f"📊 Datos recibidos para informe: {data}")
        
        # Extraer configuración
        config = {
            'titulo': data.get('titulo', 'Informe Geotecnico'),
            'cliente': data.get('cliente', ''),
            'proyecto': data.get('proyecto', ''),
            'ubicacion': data.get('ubicacion', ''),
            'fecha': data.get('fecha', datetime.now().strftime('%Y-%m-%d')),
            'incluirHumedad': data.get('incluirHumedad', False),
            'incluirAtterberg': data.get('incluirAtterberg', False),
            'incluirClasificacion': data.get('incluirClasificacion', False),
            'incluirFases': data.get('incluirFases', False),
            'incluirGraficos': data.get('incluirGraficos', False),
            'formato': data.get('formato', 'pdf')
        }
        
        print(f"📄 Generando informe en formato: {config['formato']}")
        
        # Generar según formato
        if config['formato'] == 'pdf':
            return generar_pdf(config)
        elif config['formato'] == 'excel':
            return generar_excel(config)
        elif config['formato'] == 'word':
            return generar_word(config)
        else:
            return jsonify({'success': False, 'error': 'Formato no soportado'}), 400
            
    except Exception as e:
        import traceback
        print(f"❌ Error generando informe: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


def generar_pdf(config):
    """Generar informe en formato PDF profesional"""
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter,
                              rightMargin=0.75*inch, leftMargin=0.75*inch,
                              topMargin=0.75*inch, bottomMargin=0.75*inch)
        
        # Contenedor de elementos
        elementos = []
        styles = getSampleStyleSheet()
        
        # Estilo personalizado para título
        titulo_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2d5016'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        # Estilo para subtítulos
        subtitulo_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#4a7c59'),
            spaceAfter=12,
            spaceBefore=20,
            fontName='Helvetica-Bold'
        )
        
        # Título principal
        elementos.append(Paragraph("INFORME GEOTÉCNICO", titulo_style))
        elementos.append(Paragraph(config['titulo'], styles['Heading2']))
        elementos.append(Spacer(1, 0.3*inch))
        
        # Información del proyecto en tabla
        info_data = [
            ['Campo', 'Información'],
            ['Fecha del Informe:', config['fecha']],
        ]
        
        if config['cliente']:
            info_data.append(['Cliente:', config['cliente']])
        if config['proyecto']:
            info_data.append(['Proyecto:', config['proyecto']])
        if config['ubicacion']:
            info_data.append(['Ubicación:', config['ubicacion']])
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a7c59')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        
        elementos.append(info_table)
        elementos.append(Spacer(1, 0.3*inch))
        
        # Ensayos incluidos
        elementos.append(Paragraph("ENSAYOS INCLUIDOS", subtitulo_style))
        
        ensayos_lista = []
        if config['incluirHumedad']:
            ensayos_lista.append('• Contenido de Humedad')
        if config['incluirAtterberg']:
            ensayos_lista.append('• Límites de Atterberg')
        if config['incluirClasificacion']:
            ensayos_lista.append('• Clasificación de Suelos (AASHTO)')
        if config['incluirFases']:
            ensayos_lista.append('• Fases del Suelo')
        
        for ensayo in ensayos_lista:
            elementos.append(Paragraph(ensayo, styles['Normal']))
        
        elementos.append(Spacer(1, 0.2*inch))
        
        # Agregar datos de ensayos
        if config['incluirHumedad']:
            elementos.extend(agregar_datos_humedad_pdf_mejorado(config))
        
        if config['incluirAtterberg']:
            elementos.extend(agregar_datos_atterberg_pdf_mejorado(config))
        
        if config['incluirClasificacion']:
            elementos.extend(agregar_datos_clasificacion_pdf_mejorado(config))
        
        # Pie de página
        elementos.append(Spacer(1, 0.5*inch))
        pie_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        elementos.append(Paragraph(
            f"Generado por Sistema HSGVA - {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            pie_style
        ))
        
        # Construir PDF
        doc.build(elementos)
        buffer.seek(0)
        
        filename = f"informe_{config['fecha']}.pdf"
        print(f"✅ PDF generado: {filename}")
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        import traceback
        print(f"❌ Error en generar_pdf: {str(e)}")
        print(traceback.format_exc())
        raise


def generar_excel(config):
    """Generar informe en formato Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Informe Geotecnico"
    
    # Estilos
    header_fill = PatternFill(start_color="4A7C59", end_color="4A7C59", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    title_font = Font(bold=True, size=16)
    
    # Título
    ws.merge_cells('A1:E1')
    ws['A1'] = "INFORME GEOTECNICO"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Información del proyecto
    row = 3
    ws[f'A{row}'] = "Titulo:"
    ws[f'B{row}'] = config['titulo']
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    if config['cliente']:
        ws[f'A{row}'] = "Cliente:"
        ws[f'B{row}'] = config['cliente']
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    if config['proyecto']:
        ws[f'A{row}'] = "Proyecto:"
        ws[f'B{row}'] = config['proyecto']
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    if config['ubicacion']:
        ws[f'A{row}'] = "Ubicacion:"
        ws[f'B{row}'] = config['ubicacion']
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    ws[f'A{row}'] = "Fecha:"
    ws[f'B{row}'] = config['fecha']
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 2
    
    # Datos de ensayos
    if config['incluirHumedad']:
        row = agregar_datos_humedad_excel(ws, row, config)
    
    if config['incluirAtterberg']:
        row = agregar_datos_atterberg_excel(ws, row, config)
    
    if config['incluirClasificacion']:
        row = agregar_datos_clasificacion_excel(ws, row, config)
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"informe_{config['fecha']}.xlsx"
    print(f"✅ Excel generado: {filename}")
    
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def generar_word(config):
    """Generar informe en formato Word (simplificado como TXT por ahora)"""
    contenido = f"""
INFORME GEOTECNICO
{'='*60}

{config['titulo']}

INFORMACION DEL PROYECTO
{'='*60}
Cliente: {config['cliente']}
Proyecto: {config['proyecto']}
Ubicacion: {config['ubicacion']}
Fecha: {config['fecha']}

ENSAYOS INCLUIDOS
{'='*60}
"""
    
    if config['incluirHumedad']:
        contenido += "\n- Contenido de Humedad"
    if config['incluirAtterberg']:
        contenido += "\n- Limites de Atterberg"
    if config['incluirClasificacion']:
        contenido += "\n- Clasificacion de Suelos (AASHTO)"
    if config['incluirFases']:
        contenido += "\n- Fases del Suelo"
    
    contenido += f"\n\n{'='*60}\nGenerado por Sistema HSGVA - {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
    
    buffer = io.BytesIO()
    buffer.write(contenido.encode('utf-8'))
    buffer.seek(0)
    
    filename = f"informe_{config['fecha']}.txt"
    print(f"✅ TXT generado: {filename}")
    
    return send_file(
        buffer,
        mimetype='text/plain',
        as_attachment=True,
        download_name=filename
    )


def agregar_datos_humedad_pdf_mejorado(config):
    """Agregar datos de humedad con tabla y gráfico al PDF"""
    elementos = []
    styles = getSampleStyleSheet()
    
    try:
        datos = hidrometria_etl.load_data()
        
        subtitulo_style = ParagraphStyle(
            'Subtitulo',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#4a7c59'),
            spaceAfter=12,
            spaceBefore=20,
            fontName='Helvetica-Bold'
        )
        
        elementos.append(Paragraph("CONTENIDO DE HUMEDAD", subtitulo_style))
        elementos.append(Spacer(1, 0.1*inch))
        
        # Resumen estadístico
        if len(datos) > 0:
            humedades = [d.get('Humedad (%)', 0) for d in datos]
            promedio = sum(humedades) / len(humedades)
            maximo = max(humedades)
            minimo = min(humedades)
            
            resumen_text = f"<b>Total de muestras:</b> {len(datos)} | <b>Promedio:</b> {promedio:.2f}% | <b>Máximo:</b> {maximo:.2f}% | <b>Mínimo:</b> {minimo:.2f}%"
            elementos.append(Paragraph(resumen_text, styles['Normal']))
            elementos.append(Spacer(1, 0.15*inch))
            
            # Tabla de datos (primeras 10 muestras)
            tabla_data = [['N°', 'Humedad (%)', 'Temperatura (°C)', 'Peso Húmedo (g)', 'Peso Seco (g)']]
            
            for i, dato in enumerate(datos[:10], 1):
                tabla_data.append([
                    str(i),
                    f"{dato.get('Humedad (%)', 0):.2f}",
                    f"{dato.get('Temperatura', 0):.1f}",
                    f"{dato.get('Peso Humedo', 0):.2f}",
                    f"{dato.get('Peso Seco', 0):.2f}"
                ])
            
            tabla = Table(tabla_data, colWidths=[0.5*inch, 1.2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a7c59')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            elementos.append(tabla)
            
            # Generar gráfico si está habilitado
            if config.get('incluirGraficos'):
                elementos.append(Spacer(1, 0.2*inch))
                grafico_buffer = generar_grafico_humedad(datos[:10])
                if grafico_buffer:
                    img = Image(grafico_buffer, width=5*inch, height=3*inch)
                    elementos.append(img)
        
        elementos.append(Spacer(1, 0.2*inch))
        
    except Exception as e:
        print(f"Error agregando humedad: {e}")
        elementos.append(Paragraph(f"Error al cargar datos de humedad", styles['Normal']))
    
    return elementos


def agregar_datos_atterberg_pdf_mejorado(config):
    """Agregar datos de Atterberg con tabla al PDF"""
    elementos = []
    styles = getSampleStyleSheet()
    
    try:
        datos = atterberg_etl.load_data()
        
        subtitulo_style = ParagraphStyle(
            'Subtitulo',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#4a7c59'),
            spaceAfter=12,
            spaceBefore=20,
            fontName='Helvetica-Bold'
        )
        
        elementos.append(Paragraph("LÍMITES DE ATTERBERG", subtitulo_style))
        elementos.append(Spacer(1, 0.1*inch))
        
        if len(datos) > 0:
            elementos.append(Paragraph(f"<b>Total de ensayos:</b> {len(datos)}", styles['Normal']))
            elementos.append(Spacer(1, 0.15*inch))
            
            # Tabla de datos
            tabla_data = [['N°', 'Límite Líquido (%)', 'Límite Plástico (%)', 'Índice Plasticidad']]
            
            for i, dato in enumerate(datos, 1):
                ll = dato.get('Limite Liquido', 0)
                lp = dato.get('Limite Plastico', 0)
                ip = ll - lp if ll and lp else 0
                
                tabla_data.append([
                    str(i),
                    f"{ll:.2f}" if ll else 'N/A',
                    f"{lp:.2f}" if lp else 'N/A',
                    f"{ip:.2f}"
                ])
            
            tabla = Table(tabla_data, colWidths=[0.8*inch, 1.8*inch, 1.8*inch, 1.8*inch])
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            elementos.append(tabla)
            
            # Generar gráfico si está habilitado
            if config.get('incluirGraficos'):
                elementos.append(Spacer(1, 0.2*inch))
                grafico_buffer = generar_grafico_atterberg(datos)
                if grafico_buffer:
                    img = Image(grafico_buffer, width=5*inch, height=3*inch)
                    elementos.append(img)
        
        elementos.append(Spacer(1, 0.2*inch))
        
    except Exception as e:
        print(f"Error agregando Atterberg: {e}")
        elementos.append(Paragraph(f"Error al cargar datos de Atterberg", styles['Normal']))
    
    return elementos


def agregar_datos_clasificacion_pdf_mejorado(config):
    """Agregar datos de clasificación con tabla al PDF"""
    elementos = []
    styles = getSampleStyleSheet()
    
    try:
        datos = clasificacion_etl.load_data()
        
        subtitulo_style = ParagraphStyle(
            'Subtitulo',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#4a7c59'),
            spaceAfter=12,
            spaceBefore=20,
            fontName='Helvetica-Bold'
        )
        
        elementos.append(Paragraph("CLASIFICACIÓN DE SUELOS (AASHTO)", subtitulo_style))
        elementos.append(Spacer(1, 0.1*inch))
        
        if len(datos) > 0:
            elementos.append(Paragraph(f"<b>Total de muestras:</b> {len(datos)}", styles['Normal']))
            elementos.append(Spacer(1, 0.15*inch))
            
            # Tabla de datos
            tabla_data = [['N°', 'Clasificación AASHTO', '% Pasa Tamiz #200', 'Descripción']]
            
            for i, dato in enumerate(datos, 1):
                tabla_data.append([
                    str(i),
                    dato.get('Clasificacion AASHTO', 'N/A'),
                    f"{dato.get('Porcentaje que Pasa', 0):.2f}%",
                    dato.get('Descripcion', 'N/A')
                ])
            
            tabla = Table(tabla_data, colWidths=[0.5*inch, 1.8*inch, 1.8*inch, 2.1*inch])
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            elementos.append(tabla)
        
        elementos.append(Spacer(1, 0.2*inch))
        
    except Exception as e:
        print(f"Error agregando clasificación: {e}")
        elementos.append(Paragraph(f"Error al cargar datos de clasificación", styles['Normal']))
    
    return elementos


def generar_grafico_humedad(datos):
    """Generar gráfico de barras para contenido de humedad"""
    try:
        plt.figure(figsize=(6, 3.5))
        
        muestras = [f"M{i+1}" for i in range(len(datos))]
        humedades = [d.get('Humedad (%)', 0) for d in datos]
        
        plt.bar(muestras, humedades, color='#3b82f6', alpha=0.7, edgecolor='navy')
        plt.xlabel('Muestras', fontsize=10)
        plt.ylabel('Humedad (%)', fontsize=10)
        plt.title('Contenido de Humedad por Muestra', fontsize=12, fontweight='bold')
        plt.xticks(rotation=45, ha='right', fontsize=8)
        plt.yticks(fontsize=8)
        plt.grid(axis='y', alpha=0.3, linestyle='--')
        plt.tight_layout()
        
        # Guardar en buffer
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
        plt.close()
        buffer.seek(0)
        
        return buffer
    except Exception as e:
        print(f"Error generando gráfico de humedad: {e}")
        return None


def generar_grafico_atterberg(datos):
    """Generar gráfico comparativo de límites de Atterberg"""
    try:
        plt.figure(figsize=(6, 3.5))
        
        ensayos = [f"E{i+1}" for i in range(len(datos))]
        ll = [d.get('Limite Liquido', 0) for d in datos]
        lp = [d.get('Limite Plastico', 0) for d in datos]
        
        x = range(len(ensayos))
        width = 0.35
        
        plt.bar([i - width/2 for i in x], ll, width, label='Límite Líquido', color='#667eea', alpha=0.7)
        plt.bar([i + width/2 for i in x], lp, width, label='Límite Plástico', color='#f59e0b', alpha=0.7)
        
        plt.xlabel('Ensayos', fontsize=10)
        plt.ylabel('Límite (%)', fontsize=10)
        plt.title('Límites de Atterberg', fontsize=12, fontweight='bold')
        plt.xticks(x, ensayos, fontsize=8)
        plt.yticks(fontsize=8)
        plt.legend(fontsize=8)
        plt.grid(axis='y', alpha=0.3, linestyle='--')
        plt.tight_layout()
        
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
        plt.close()
        buffer.seek(0)
        
        return buffer
    except Exception as e:
        print(f"Error generando gráfico de Atterberg: {e}")
        return None


def agregar_datos_humedad_pdf(c, y):
    """Agregar datos de humedad al PDF"""
    try:
        datos = hidrometria_etl.load_data()
        if y < 2*inch:
            c.showPage()
            y = 10*inch
        
        c.setFont("Helvetica-Bold", 11)
        c.drawString(1*inch, y, "CONTENIDO DE HUMEDAD")
        y -= 0.3*inch
        
        c.setFont("Helvetica", 9)
        c.drawString(1*inch, y, f"Total de muestras: {len(datos)}")
        y -= 0.25*inch
        
        if len(datos) > 0:
            promedio = sum(d.get('Humedad (%)', 0) for d in datos) / len(datos)
            c.drawString(1*inch, y, f"Humedad promedio: {promedio:.2f}%")
            y -= 0.4*inch
    except Exception as e:
        print(f"Error agregando humedad: {e}")
    
    return y


def agregar_datos_atterberg_pdf(c, y):
    """Agregar datos de Atterberg al PDF"""
    try:
        datos = atterberg_etl.load_data()
        if y < 2*inch:
            c.showPage()
            y = 10*inch
        
        c.setFont("Helvetica-Bold", 11)
        c.drawString(1*inch, y, "LIMITES DE ATTERBERG")
        y -= 0.3*inch
        
        c.setFont("Helvetica", 9)
        c.drawString(1*inch, y, f"Total de ensayos: {len(datos)}")
        y -= 0.4*inch
    except Exception as e:
        print(f"Error agregando Atterberg: {e}")
    
    return y


def agregar_datos_clasificacion_pdf(c, y):
    """Agregar datos de clasificación al PDF"""
    try:
        datos = clasificacion_etl.load_data()
        if y < 2*inch:
            c.showPage()
            y = 10*inch
        
        c.setFont("Helvetica-Bold", 11)
        c.drawString(1*inch, y, "CLASIFICACION DE SUELOS (AASHTO)")
        y -= 0.3*inch
        
        c.setFont("Helvetica", 9)
        c.drawString(1*inch, y, f"Total de muestras: {len(datos)}")
        y -= 0.4*inch
    except Exception as e:
        print(f"Error agregando clasificacion: {e}")
    
    return y


def agregar_datos_humedad_excel(ws, row, config):
    """Agregar datos de humedad al Excel con formato y gráfico"""
    try:
        datos = hidrometria_etl.load_data()
        
        # Título sección
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = "CONTENIDO DE HUMEDAD"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="2D5016")
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1
        
        # Resumen estadístico
        if len(datos) > 0:
            humedades = [d.get('Humedad (%)', 0) for d in datos]
            promedio = sum(humedades) / len(humedades)
            ws[f'A{row}'] = f"Muestras: {len(datos)}"
            ws[f'B{row}'] = f"Promedio: {promedio:.2f}%"
            ws[f'C{row}'] = f"Máximo: {max(humedades):.2f}%"
            ws[f'D{row}'] = f"Mínimo: {min(humedades):.2f}%"
            row += 1
        
        row += 1
        
        # Encabezados de tabla
        headers = ["N°", "Humedad (%)", "Temperatura (°C)", "Peso Húmedo (g)", "Peso Seco (g)"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4A7C59", end_color="4A7C59", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Datos
        start_data_row = row
        for i, dato in enumerate(datos[:10], 1):
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=dato.get('Humedad (%)', 0))
            ws.cell(row=row, column=3, value=dato.get('Temperatura', 0))
            ws.cell(row=row, column=4, value=dato.get('Peso Humedo', 0))
            ws.cell(row=row, column=5, value=dato.get('Peso Seco', 0))
            
            # Formatear números
            for col in range(2, 6):
                ws.cell(row=row, column=col).number_format = '0.00'
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            
            row += 1
        
        # Agregar gráfico si está habilitado
        if config.get('incluirGraficos') and len(datos) > 0:
            chart = BarChart()
            chart.title = "Contenido de Humedad por Muestra"
            chart.y_axis.title = "Humedad (%)"
            chart.x_axis.title = "Muestra"
            
            # Referencias de datos
            data = Reference(ws, min_col=2, min_row=start_data_row-1, max_row=row-1)
            cats = Reference(ws, min_col=1, min_row=start_data_row, max_row=row-1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 10
            chart.width = 20
            
            # Posicionar gráfico
            ws.add_chart(chart, f'G{start_data_row-2}')
        
        row += 2
    except Exception as e:
        print(f"Error agregando humedad a Excel: {e}")
    
    return row


def agregar_datos_atterberg_excel(ws, row, config):
    """Agregar datos de Atterberg al Excel con formato y gráfico"""
    try:
        datos = atterberg_etl.load_data()
        
        # Título sección
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = "LÍMITES DE ATTERBERG"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="667EEA")
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1
        
        ws[f'A{row}'] = f"Total de ensayos: {len(datos)}"
        row += 2
        
        # Encabezados
        headers = ["N°", "Límite Líquido (%)", "Límite Plástico (%)", "Índice Plasticidad"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Datos
        start_data_row = row
        for i, dato in enumerate(datos, 1):
            ll = dato.get('Limite Liquido', 0)
            lp = dato.get('Limite Plastico', 0)
            ip = ll - lp if ll and lp else 0
            
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=ll)
            ws.cell(row=row, column=3, value=lp)
            ws.cell(row=row, column=4, value=ip)
            
            for col in range(2, 5):
                ws.cell(row=row, column=col).number_format = '0.00'
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            
            row += 1
        
        # Agregar gráfico comparativo
        if config.get('incluirGraficos') and len(datos) > 0:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Límites de Atterberg"
            chart.y_axis.title = "Límite (%)"
            chart.x_axis.title = "Ensayo"
            
            data = Reference(ws, min_col=2, min_row=start_data_row-1, max_col=3, max_row=row-1)
            cats = Reference(ws, min_col=1, min_row=start_data_row, max_row=row-1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 10
            chart.width = 20
            
            ws.add_chart(chart, f'G{start_data_row-2}')
        
        row += 2
    except Exception as e:
        print(f"Error agregando Atterberg a Excel: {e}")
    
    return row


def agregar_datos_clasificacion_excel(ws, row, config):
    """Agregar datos de clasificación al Excel"""
    try:
        datos = clasificacion_etl.load_data()
        
        # Título sección
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = "CLASIFICACIÓN DE SUELOS (AASHTO)"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="10B981")
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1
        
        ws[f'A{row}'] = f"Total de muestras: {len(datos)}"
        row += 2
        
        # Encabezados
        headers = ["N°", "Clasificación AASHTO", "% Pasa Tamiz #200", "Descripción"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Datos
        for i, dato in enumerate(datos, 1):
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=dato.get('Clasificacion AASHTO', 'N/A'))
            ws.cell(row=row, column=3, value=dato.get('Porcentaje que Pasa', 0))
            ws.cell(row=row, column=4, value=dato.get('Descripcion', 'N/A'))
            
            ws.cell(row=row, column=3).number_format = '0.00'
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            
            row += 1
        
        row += 2
    except Exception as e:
        print(f"Error agregando clasificacion a Excel: {e}")
    
    return row


# Manejo de errores
@app.errorhandler(404)
def not_found(e):
    return jsonify({'error': 'Recurso no encontrado'}), 404


@app.errorhandler(500)
def server_error(e):
    return jsonify({'error': 'Error interno del servidor'}), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
